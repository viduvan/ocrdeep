#!/usr/bin/env python3
"""
Export GTGT Invoice Extraction Report to Excel.
Separate sheets for Regex and LLM results.
"""
import sys, os, re
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from tests.test_llm_extraction import run_regex_parser, run_llm_extractor, ALL_FIELDS, normalize_date
import tests.test_cases_gtgt as tc

ITEM_FIELDS = ["productName", "quantity", "unitPrice", "amount"]

# ─── Styles ──────────────────────────────────────────────────────────────────
H_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
H_FONT = Font(bold=True, color="FFFFFF", size=10)
H_FONT_SM = Font(bold=True, color="FFFFFF", size=9)
MATCH_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
MM_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
PCT_HIGH = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
PCT_MED = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
PCT_LOW = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
BORDER = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
WRAP = Alignment(wrap_text=True, vertical='top')


def get_case_ids():
    nums = set()
    for attr in dir(tc):
        m = re.match(r'rawtext_(\d+)', attr)
        if m:
            nums.add(int(m.group(1)))
    return sorted(nums)


def get_data(case_ids):
    results = []
    for i, c in enumerate(case_ids):
        raw_fn = getattr(tc, f"rawtext_{c:02d}", None) or getattr(tc, f"rawtext_{c}", None)
        zoom_fn = getattr(tc, f"zoomtext_{c:02d}", None) or getattr(tc, f"zoomtext_{c}", None)
        if not raw_fn:
            continue
        raw = raw_fn()
        zoom = zoom_fn() if zoom_fn else ""
        print(f"  [{i+1}/{len(case_ids)}] Case {c}...", end='\r')

        regex_r = run_regex_parser(raw, zoom)
        try:
            llm_r = run_llm_extractor(raw, zoom)
            if not llm_r:
                llm_r = {}
        except Exception as e:
            print(f"  Case {c}: LLM error - {e}")
            llm_r = {}

        if llm_r.get("invoiceDate"):
            llm_r["invoiceDate"] = normalize_date(llm_r["invoiceDate"])

        results.append({'case': c, 'regex': regex_r, 'llm': llm_r})
    print(f"\n  ✅ Processed {len(results)} cases.")
    return results


def smart_compare(field, r_val, l_val):
    """Smart field comparison handling known patterns."""
    if r_val is None and l_val is None:
        return 'skip'
    if r_val is not None and l_val is None:
        return 'regex_only'
    if r_val is None and l_val is not None:
        return 'llm_only'

    r_s = str(r_val).strip()
    l_s = str(l_val).strip()
    if r_s == l_s:
        return 'match'

    # Numeric comparison
    try:
        r_f, l_f = float(r_s), float(l_s)
        if r_f == l_f or abs(r_f) == abs(l_f):
            return 'match'
    except:
        pass

    # invoiceTotalInWord: fuzzy (strip punctuation, prefix noise)
    if field == 'invoiceTotalInWord':
        r_n = re.sub(r'[./,*\[\]\s]+', ' ', r_s).strip().lower()
        l_n = re.sub(r'[./,*\[\]\s]+', ' ', l_s).strip().lower()
        if r_n == l_n or r_n[:40] == l_n[:40] or l_n[:40] == r_n[:40]:
            return 'match'

    return 'mismatch'


def detect_serial_swap(regex_r, llm_r):
    """Check if regex swapped invoiceSerial <-> invoiceFormNo."""
    r_serial = str(regex_r.get("invoiceSerial", "") or "")
    r_formno = str(regex_r.get("invoiceFormNo", "") or "")
    l_serial = str(llm_r.get("invoiceSerial", "") or "")
    l_formno = str(llm_r.get("invoiceFormNo", "") or "")
    return (r_serial and r_formno and l_serial and l_formno
            and r_serial == l_formno and r_formno == l_serial)


def styled_cell(ws, row, col, value, font=None, fill=None, border=BORDER, align=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font: cell.font = font
    if fill: cell.fill = fill
    if border: cell.border = border
    if align: cell.alignment = align
    return cell


def create_excel(results, output_path):
    wb = Workbook()

    # ═══════════════════════════════════════════════════════════════════════════
    # Sheet 1: SUMMARY
    # ═══════════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_properties.tabColor = "1F4E79"

    ws.merge_cells('A1:G1')
    ws['A1'] = "BÁO CÁO ĐÁNH GIÁ LLM EXTRACTOR - HÓA ĐƠN GTGT"
    ws['A1'].font = Font(bold=True, size=14, color="1F4E79")
    ws['A3'] = f"Tổng số cases: {len(results)} | Model: Qwen3-32B (FPT Cloud)"

    # Calculate per-field stats WITH smart comparison + serial swap handling
    field_stats = {f: {'match': 0, 'mismatch': 0, 'compared': 0,
                       'regex_only': 0, 'llm_only': 0, 'skip': 0} for f in ALL_FIELDS}

    for r in results:
        swapped = detect_serial_swap(r['regex'], r['llm'])
        for field in ALL_FIELDS:
            rv = r['regex'].get(field)
            lv = r['llm'].get(field)
            status = smart_compare(field, rv, lv)

            # Handle serial/formNo swap
            if swapped and field in ('invoiceSerial', 'invoiceFormNo') and status == 'mismatch':
                status = 'match'  # It's a known regex swap, not a real mismatch

            field_stats[field][status] += 1
            if status in ('match', 'mismatch'):
                field_stats[field]['compared'] += 1

    # Summary table: per-field accuracy
    row = 5
    headers = ['Trường (Field)', 'Khớp (Match)', 'Sai lệch (Mismatch)', 'Tỉ lệ đúng (%)',
               'Chỉ Regex', 'Chỉ LLM', 'Ghi chú']
    for c, h in enumerate(headers, 1):
        styled_cell(ws, row, c, h, H_FONT, H_FILL)
    ws.row_dimensions[row].height = 20

    for i, field in enumerate(ALL_FIELDS):
        r = row + 1 + i
        s = field_stats[field]
        if s['compared'] > 0:
            pct = s['match'] / s['compared'] * 100
            pct_str = f"{pct:.1f}%"
        else:
            pct = -1
            pct_str = "N/A"

        styled_cell(ws, r, 1, field, Font(bold=True, size=10))
        styled_cell(ws, r, 2, s['match'])
        mm_cell = styled_cell(ws, r, 3, s['mismatch'])
        pct_cell = styled_cell(ws, r, 4, pct_str, Font(bold=True))
        styled_cell(ws, r, 5, s['regex_only'])
        styled_cell(ws, r, 6, s['llm_only'])

        # Color code percentage
        if pct >= 95:
            pct_cell.fill = PCT_HIGH
        elif pct >= 80:
            pct_cell.fill = PCT_MED
        elif pct >= 0:
            pct_cell.fill = PCT_LOW

        if s['mismatch'] > 0:
            mm_cell.fill = MM_FILL

        # Auto-note
        note = ""
        if field in ('invoiceSerial', 'invoiceFormNo'):
            note = "Regex thường swap Serial↔FormNo"
        elif field == 'invoiceTotalInWord':
            note = "So sánh fuzzy (dấu chấm, dấu *)"
        elif field == 'taxPercent':
            note = "Regex thường trả 0%"
        styled_cell(ws, r, 7, note, Font(italic=True, color="666666", size=9))

    # Overall summary row
    total_row = row + len(ALL_FIELDS) + 2
    total_m = sum(s['match'] for s in field_stats.values())
    total_mm = sum(s['mismatch'] for s in field_stats.values())
    total_c = total_m + total_mm
    styled_cell(ws, total_row, 1, "TỔNG CỘNG", Font(bold=True, size=12, color="1F4E79"))
    styled_cell(ws, total_row, 2, total_m, Font(bold=True, size=12))
    styled_cell(ws, total_row, 3, total_mm, Font(bold=True, size=12))
    pct_total = f"{total_m/total_c*100:.1f}%" if total_c else "N/A"
    styled_cell(ws, total_row, 4, pct_total, Font(bold=True, size=12, color="1F4E79"))

    # Column widths
    widths = {'A': 26, 'B': 14, 'C': 16, 'D': 16, 'E': 12, 'F': 12, 'G': 35}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # ═══════════════════════════════════════════════════════════════════════════
    # Sheet 2: REGEX Fields
    # ═══════════════════════════════════════════════════════════════════════════
    ws_regex = wb.create_sheet("Regex Fields")
    ws_regex.sheet_properties.tabColor = "BDD7EE"

    headers = ['Case'] + ALL_FIELDS
    for c, h in enumerate(headers, 1):
        styled_cell(ws_regex, 1, c, h, H_FONT_SM, H_FILL)

    for row_idx, r in enumerate(results, 2):
        styled_cell(ws_regex, row_idx, 1, r['case'])
        for col_idx, field in enumerate(ALL_FIELDS, 2):
            val = r['regex'].get(field)
            val_str = str(val)[:60] if val is not None else ''
            styled_cell(ws_regex, row_idx, col_idx, val_str, Font(size=8))

    ws_regex.freeze_panes = 'B2'
    ws_regex.column_dimensions['A'].width = 6

    # ═══════════════════════════════════════════════════════════════════════════
    # Sheet 3: LLM Fields
    # ═══════════════════════════════════════════════════════════════════════════
    ws_llm = wb.create_sheet("LLM Fields")
    ws_llm.sheet_properties.tabColor = "E2EFDA"

    for c, h in enumerate(headers, 1):
        styled_cell(ws_llm, 1, c, h, H_FONT_SM, H_FILL)

    for row_idx, r in enumerate(results, 2):
        styled_cell(ws_llm, row_idx, 1, r['case'])
        for col_idx, field in enumerate(ALL_FIELDS, 2):
            val = r['llm'].get(field)
            val_str = str(val)[:60] if val is not None else ''
            styled_cell(ws_llm, row_idx, col_idx, val_str, Font(size=8))

    ws_llm.freeze_panes = 'B2'
    ws_llm.column_dimensions['A'].width = 6

    # ═══════════════════════════════════════════════════════════════════════════
    # Sheet 4: Field Comparison (match/mismatch highlights)
    # ═══════════════════════════════════════════════════════════════════════════
    ws_cmp = wb.create_sheet("Field Comparison")
    ws_cmp.sheet_properties.tabColor = "2E75B6"

    cmp_headers = ['Case']
    for f in ALL_FIELDS:
        cmp_headers.extend([f"{f}_Regex", f"{f}_LLM", f"{f}_Status"])
    for c, h in enumerate(cmp_headers, 1):
        styled_cell(ws_cmp, 1, c, h, Font(bold=True, color="FFFFFF", size=7), H_FILL,
                    align=Alignment(wrap_text=True, text_rotation=90))

    for row_idx, r in enumerate(results, 2):
        swapped = detect_serial_swap(r['regex'], r['llm'])
        styled_cell(ws_cmp, row_idx, 1, r['case'])
        col = 2
        for field in ALL_FIELDS:
            rv = r['regex'].get(field)
            lv = r['llm'].get(field)
            status = smart_compare(field, rv, lv)
            if swapped and field in ('invoiceSerial', 'invoiceFormNo') and status == 'mismatch':
                status = 'match'

            r_str = str(rv)[:50] if rv is not None else ''
            l_str = str(lv)[:50] if lv is not None else ''
            icon = '✅' if status == 'match' else ('❌' if status == 'mismatch' else '')

            cell_r = styled_cell(ws_cmp, row_idx, col, r_str, Font(size=7))
            cell_l = styled_cell(ws_cmp, row_idx, col+1, l_str, Font(size=7))
            cell_s = styled_cell(ws_cmp, row_idx, col+2, icon, Font(size=8))

            if status == 'mismatch':
                cell_r.fill = MM_FILL; cell_l.fill = MM_FILL; cell_s.fill = MM_FILL
            elif status == 'match':
                cell_s.fill = MATCH_FILL
            col += 3

    ws_cmp.freeze_panes = 'B2'
    ws_cmp.column_dimensions['A'].width = 6

    # ═══════════════════════════════════════════════════════════════════════════
    # Sheet 5: Regex Items
    # ═══════════════════════════════════════════════════════════════════════════
    ws_ri = wb.create_sheet("Regex Items")
    ws_ri.sheet_properties.tabColor = "BDD7EE"
    item_h = ['Case', 'Item#', 'productName', 'quantity', 'unitPrice', 'amount']
    for c, h in enumerate(item_h, 1):
        styled_cell(ws_ri, 1, c, h, H_FONT_SM, H_FILL)

    row = 2
    for r in results:
        items = r['regex'].get('itemList', []) or []
        for i, item in enumerate(items):
            styled_cell(ws_ri, row, 1, r['case'])
            styled_cell(ws_ri, row, 2, i+1)
            for j, f in enumerate(ITEM_FIELDS):
                v = item.get(f)
                styled_cell(ws_ri, row, 3+j, str(v)[:50] if v is not None else '', Font(size=8))
            row += 1

    ws_ri.freeze_panes = 'C2'
    ws_ri.column_dimensions['A'].width = 6; ws_ri.column_dimensions['B'].width = 6
    ws_ri.column_dimensions['C'].width = 35; ws_ri.column_dimensions['D'].width = 12
    ws_ri.column_dimensions['E'].width = 14; ws_ri.column_dimensions['F'].width = 16

    # ═══════════════════════════════════════════════════════════════════════════
    # Sheet 6: LLM Items
    # ═══════════════════════════════════════════════════════════════════════════
    ws_li = wb.create_sheet("LLM Items")
    ws_li.sheet_properties.tabColor = "E2EFDA"
    for c, h in enumerate(item_h, 1):
        styled_cell(ws_li, 1, c, h, H_FONT_SM, H_FILL)

    row = 2
    for r in results:
        items = r['llm'].get('itemList', []) or []
        for i, item in enumerate(items):
            styled_cell(ws_li, row, 1, r['case'])
            styled_cell(ws_li, row, 2, i+1)
            for j, f in enumerate(ITEM_FIELDS):
                v = item.get(f)
                styled_cell(ws_li, row, 3+j, str(v)[:50] if v is not None else '', Font(size=8))
            row += 1

    ws_li.freeze_panes = 'C2'
    ws_li.column_dimensions['A'].width = 6; ws_li.column_dimensions['B'].width = 6
    ws_li.column_dimensions['C'].width = 35; ws_li.column_dimensions['D'].width = 12
    ws_li.column_dimensions['E'].width = 14; ws_li.column_dimensions['F'].width = 16

    # ═══════════════════════════════════════════════════════════════════════════
    wb.save(output_path)
    print(f"\n✅ Report saved: {output_path}")
    print(f"   6 Sheets: Summary, Regex Fields, LLM Fields, Field Comparison, Regex Items, LLM Items")
    print(f"   Cases: {len(results)}")


if __name__ == "__main__":
    output = os.path.join(os.path.dirname(__file__), '..', 'gtgt_llm_report.xlsx')
    print("Generating GTGT LLM Extraction Report...")
    print("=" * 60)
    case_ids = get_case_ids()
    print(f"  Found {len(case_ids)} test cases")
    results = get_data(case_ids)
    create_excel(results, output)
