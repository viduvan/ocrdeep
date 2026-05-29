#!/usr/bin/env python3
"""
Export LLM-only Accuracy Report to Excel.
Evaluates LLM extraction independently using internal consistency checks.
No regex comparison - pure LLM self-validation.
"""
import sys, os, re, argparse
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from tests.test_llm_extraction import run_llm_extractor, normalize_date

tc = None  # set dynamically

ALL_FIELDS = [
    "invoiceName", "invoiceDate", "invoiceID", "invoiceSerial", "invoiceFormNo",
    "sellerName", "sellerTaxCode", "sellerAddress", "sellerPhoneNumber",
    "sellerBank", "sellerBankAccountNumber",
    "buyerName", "buyerTaxCode", "buyerAddress",
    "buyerBank", "buyerBankAccountNumber",
    "paymentMethod", "currency",
    "preTaxPrice", "taxPercent", "taxAmount", "totalAmount",
    "invoiceTotalInWord",
]

ITEM_FIELDS = ["productName", "quantity", "unitPrice", "amount"]

# ─── Styles ──────────────────────────────────────────────────────────────────
H_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
H2_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
H_FONT = Font(bold=True, color="FFFFFF", size=10)
H_FONT_SM = Font(bold=True, color="FFFFFF", size=9)
OK_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
WARN_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
ERR_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
BORDER = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))


def sc(ws, r, c, v, font=None, fill=None, border=BORDER, align=None):
    """Styled cell helper."""
    cell = ws.cell(row=r, column=c, value=v)
    if font: cell.font = font
    if fill: cell.fill = fill
    if border: cell.border = border
    if align: cell.alignment = align
    return cell


def get_case_ids(tc_mod):
    nums = set()
    for attr in dir(tc_mod):
        m = re.match(r'rawtext_(\d+)', attr)
        if m: nums.add(int(m.group(1)))
    return sorted(nums)


def get_llm_data(tc_mod, case_ids):
    results = []
    for i, c in enumerate(case_ids):
        raw_fn = getattr(tc_mod, f"rawtext_{c:02d}", None) or getattr(tc_mod, f"rawtext_{c}", None)
        zoom_fn = getattr(tc_mod, f"zoomtext_{c:02d}", None) or getattr(tc_mod, f"zoomtext_{c}", None)
        if not raw_fn: continue
        raw = raw_fn(); zoom = zoom_fn() if zoom_fn else ""
        print(f"  [{i+1}/{len(case_ids)}] Case {c}...", end='\r')
        try:
            r = run_llm_extractor(raw, zoom)
            if not r: r = {}
        except Exception as e:
            print(f"  Case {c}: error - {e}")
            r = {}
        if r.get("invoiceDate"):
            r["invoiceDate"] = normalize_date(r["invoiceDate"])
        results.append({'case': c, 'llm': r, 'raw': raw})
    print(f"\n  ✅ Processed {len(results)} cases.")
    return results


def safe_float(v):
    if v is None: return None
    try: return float(v)
    except: return None


def validate_case(r):
    """Run all internal consistency checks on a single LLM result."""
    llm = r['llm']
    raw = r['raw']
    checks = {}

    # ── 1. Field coverage (non-null) ──
    for f in ALL_FIELDS:
        checks[f'has_{f}'] = llm.get(f) is not None

    # ── 2. Math: totalAmount = preTaxPrice + taxAmount ──
    total = safe_float(llm.get('totalAmount'))
    pretax = safe_float(llm.get('preTaxPrice'))
    tax = safe_float(llm.get('taxAmount'))

    if total is not None and pretax is not None and tax is not None:
        expected = pretax + tax
        if abs(total - expected) < 1:
            checks['total_math'] = 'exact'
        elif abs(total) > 0 and abs(total - expected) / abs(total) < 0.01:
            checks['total_math'] = 'close'
        else:
            checks['total_math'] = 'wrong'
        checks['total_math_detail'] = f"{total:.0f} vs {expected:.0f} (diff={total-expected:.0f})"
    else:
        checks['total_math'] = 'skip'
        checks['total_math_detail'] = 'N/A (missing fields)'

    # ── 3. Math: taxAmount = preTaxPrice × taxPercent ──
    tax_pct_str = str(llm.get('taxPercent', '') or '')
    tax_pct_val = None
    pct_match = re.search(r'(\d+)', tax_pct_str)
    if pct_match:
        tax_pct_val = float(pct_match.group(1))

    if pretax is not None and tax is not None and tax_pct_val is not None and tax_pct_val > 0:
        expected_tax = pretax * tax_pct_val / 100
        if abs(tax - expected_tax) < 1:
            checks['tax_math'] = 'exact'
        elif abs(expected_tax) > 0 and abs(tax - expected_tax) / abs(expected_tax) < 0.02:
            checks['tax_math'] = 'close'
        else:
            checks['tax_math'] = 'wrong'
        checks['tax_math_detail'] = f"tax={tax:.0f} vs pretax×{tax_pct_val:.0f}%={expected_tax:.0f}"
    elif tax_pct_str in ('KCT', '0%', '0', 'KBHCTT'):
        if tax is not None and abs(safe_float(tax) or 0) < 1:
            checks['tax_math'] = 'exact'
        else:
            checks['tax_math'] = 'wrong'
        checks['tax_math_detail'] = f"KCT/0%: tax should be 0, got {tax}"
    else:
        checks['tax_math'] = 'skip'
        checks['tax_math_detail'] = 'N/A'

    # ── 4. Item math: quantity × unitPrice ≈ amount ──
    items = llm.get('itemList', []) or []
    item_ok = 0
    item_wrong = 0
    item_skip = 0
    for item in items:
        qty = safe_float(item.get('quantity'))
        price = safe_float(item.get('unitPrice'))
        amt = safe_float(item.get('amount'))
        if qty is not None and price is not None and amt is not None:
            expected = qty * price
            if abs(amt - expected) < 1:
                item_ok += 1
            elif abs(amt) > 0 and abs(amt - expected) / abs(amt) < 0.01:
                item_ok += 1
            else:
                item_wrong += 1
        else:
            item_skip += 1

    checks['item_total'] = len(items)
    checks['item_math_ok'] = item_ok
    checks['item_math_wrong'] = item_wrong
    checks['item_math_skip'] = item_skip

    # ── 5. Cross-check: invoiceTotalInWord vs totalAmount range ──
    inword = str(llm.get('invoiceTotalInWord', '') or '').lower()
    if total is not None and inword:
        # Range check — Vietnamese + English number words
        word_magnitude = None
        # Vietnamese
        if 'tỷ' in inword or 'tỉ' in inword or 'billion' in inword:
            word_magnitude = 1_000_000_000
        elif 'triệu' in inword or 'million' in inword:
            word_magnitude = 1_000_000
        elif 'nghìn' in inword or 'ngàn' in inword or 'thousand' in inword:
            word_magnitude = 1_000
        elif 'hundred' in inword or 'trăm' in inword:
            word_magnitude = 100

        if word_magnitude:
            if abs(total) >= word_magnitude * 0.5:
                checks['inword_range'] = 'match'
            else:
                checks['inword_range'] = 'mismatch'
        else:
            checks['inword_range'] = 'skip'
    else:
        checks['inword_range'] = 'skip'

    return checks


def create_excel(results, output_path, invoice_type='gtgt'):
    wb = Workbook()

    # ── Validate all cases ──
    all_checks = []
    for r in results:
        checks = validate_case(r)
        all_checks.append({'case': r['case'], 'llm': r['llm'], 'checks': checks})

    # ═══════════════════════════════════════════════════════════════════════════
    # Sheet 1: SUMMARY — Tỉ lệ phần trăm LLM
    # ═══════════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_properties.tabColor = "1F4E79"

    ws.merge_cells('A1:E1')
    ws['A1'] = f"ĐÁNH GIÁ ĐỘC LẬP LLM EXTRACTOR — {invoice_type.upper()}"
    ws['A1'].font = Font(bold=True, size=14, color="1F4E79")
    ws['A3'] = f"Tổng cases: {len(results)} | Model: Qwen3-32B (FPT Cloud)"
    ws['A3'].font = Font(size=11)

    # ── A. Field Coverage (tỉ lệ trả giá trị non-null) ──
    row = 5
    sc(ws, row, 1, "A. TỈ LỆ TRÍCH XUẤT TRƯỜNG (Coverage)", Font(bold=True, size=12, color="1F4E79"))
    row += 1
    for c, h in enumerate(['Trường', 'Có giá trị', 'Không có', 'Tỉ lệ (%)'], 1):
        sc(ws, row, c, h, H_FONT, H_FILL)

    for i, f in enumerate(ALL_FIELDS):
        r = row + 1 + i
        has = sum(1 for ac in all_checks if ac['checks'].get(f'has_{f}', False))
        no = len(all_checks) - has
        pct = has / len(all_checks) * 100
        sc(ws, r, 1, f, Font(bold=True, size=9))
        sc(ws, r, 2, has)
        sc(ws, r, 3, no)
        pct_cell = sc(ws, r, 4, f"{pct:.1f}%", Font(bold=True))
        if pct >= 95: pct_cell.fill = OK_FILL
        elif pct >= 70: pct_cell.fill = WARN_FILL
        else: pct_cell.fill = ERR_FILL

    # ── B. Math Validation ──
    row = row + len(ALL_FIELDS) + 3
    sc(ws, row, 1, "B. KIỂM TRA TÍNH TOÁN (Math Validation)", Font(bold=True, size=12, color="1F4E79"))
    row += 1
    for c, h in enumerate(['Kiểm tra', 'Đúng', 'Gần đúng', 'Sai', 'N/A', 'Tỉ lệ đúng (%)'], 1):
        sc(ws, row, c, h, H_FONT, H_FILL)

    # total = pretax + tax
    t_exact = sum(1 for ac in all_checks if ac['checks']['total_math'] == 'exact')
    t_close = sum(1 for ac in all_checks if ac['checks']['total_math'] == 'close')
    t_wrong = sum(1 for ac in all_checks if ac['checks']['total_math'] == 'wrong')
    t_skip = sum(1 for ac in all_checks if ac['checks']['total_math'] == 'skip')
    t_verifiable = t_exact + t_close + t_wrong
    t_pct = (t_exact + t_close) / t_verifiable * 100 if t_verifiable else 0

    r = row + 1
    sc(ws, r, 1, "totalAmount = preTaxPrice + taxAmount", Font(bold=True, size=9))
    sc(ws, r, 2, t_exact)
    sc(ws, r, 3, t_close)
    sc(ws, r, 4, t_wrong); ws.cell(row=r, column=4).fill = ERR_FILL if t_wrong > 0 else OK_FILL
    sc(ws, r, 5, t_skip)
    pct_cell = sc(ws, r, 6, f"{t_pct:.1f}%", Font(bold=True, size=11))
    if t_pct >= 95: pct_cell.fill = OK_FILL
    elif t_pct >= 80: pct_cell.fill = WARN_FILL
    else: pct_cell.fill = ERR_FILL

    # tax = pretax × pct
    tx_exact = sum(1 for ac in all_checks if ac['checks']['tax_math'] == 'exact')
    tx_close = sum(1 for ac in all_checks if ac['checks']['tax_math'] == 'close')
    tx_wrong = sum(1 for ac in all_checks if ac['checks']['tax_math'] == 'wrong')
    tx_skip = sum(1 for ac in all_checks if ac['checks']['tax_math'] == 'skip')
    tx_v = tx_exact + tx_close + tx_wrong
    tx_pct = (tx_exact + tx_close) / tx_v * 100 if tx_v else 0

    r += 1
    sc(ws, r, 1, "taxAmount = preTaxPrice × taxPercent", Font(bold=True, size=9))
    sc(ws, r, 2, tx_exact)
    sc(ws, r, 3, tx_close)
    sc(ws, r, 4, tx_wrong); ws.cell(row=r, column=4).fill = ERR_FILL if tx_wrong > 0 else OK_FILL
    sc(ws, r, 5, tx_skip)
    pct_cell = sc(ws, r, 6, f"{tx_pct:.1f}%", Font(bold=True, size=11))
    if tx_pct >= 95: pct_cell.fill = OK_FILL
    elif tx_pct >= 80: pct_cell.fill = WARN_FILL
    else: pct_cell.fill = ERR_FILL

    # Item math
    total_item_ok = sum(ac['checks']['item_math_ok'] for ac in all_checks)
    total_item_wrong = sum(ac['checks']['item_math_wrong'] for ac in all_checks)
    total_item_skip = sum(ac['checks']['item_math_skip'] for ac in all_checks)
    item_v = total_item_ok + total_item_wrong
    item_pct = total_item_ok / item_v * 100 if item_v else 0

    r += 1
    sc(ws, r, 1, "Item: quantity × unitPrice = amount", Font(bold=True, size=9))
    sc(ws, r, 2, total_item_ok)
    sc(ws, r, 3, 0)
    sc(ws, r, 4, total_item_wrong); ws.cell(row=r, column=4).fill = ERR_FILL if total_item_wrong > 0 else OK_FILL
    sc(ws, r, 5, total_item_skip)
    pct_cell = sc(ws, r, 6, f"{item_pct:.1f}%", Font(bold=True, size=11))
    if item_pct >= 95: pct_cell.fill = OK_FILL
    elif item_pct >= 80: pct_cell.fill = WARN_FILL
    else: pct_cell.fill = ERR_FILL

    # InWord range check
    iw_match = sum(1 for ac in all_checks if ac['checks']['inword_range'] == 'match')
    iw_mm = sum(1 for ac in all_checks if ac['checks']['inword_range'] == 'mismatch')
    iw_skip = sum(1 for ac in all_checks if ac['checks']['inword_range'] == 'skip')
    iw_v = iw_match + iw_mm
    iw_pct = iw_match / iw_v * 100 if iw_v else 0

    r += 1
    sc(ws, r, 1, "invoiceTotalInWord ~ totalAmount (range)", Font(bold=True, size=9))
    sc(ws, r, 2, iw_match)
    sc(ws, r, 3, 0)
    sc(ws, r, 4, iw_mm); ws.cell(row=r, column=4).fill = ERR_FILL if iw_mm > 0 else OK_FILL
    sc(ws, r, 5, iw_skip)
    pct_cell = sc(ws, r, 6, f"{iw_pct:.1f}%", Font(bold=True, size=11))
    if iw_pct >= 95: pct_cell.fill = OK_FILL
    elif iw_pct >= 80: pct_cell.fill = WARN_FILL
    else: pct_cell.fill = ERR_FILL

    ws.column_dimensions['A'].width = 42
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 16

    # ═══════════════════════════════════════════════════════════════════════════
    # Sheet 2: Per-Case Validation Detail
    # ═══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Case Validation")
    ws2.sheet_properties.tabColor = "2E75B6"

    headers = ['Case', 'total=pre+tax', 'Detail', 'tax=pre×pct', 'Detail',
               'Items Total', 'Items OK', 'Items Wrong', 'InWord Range',
               'totalAmount', 'preTaxPrice', 'taxAmount', 'taxPercent']
    for c, h in enumerate(headers, 1):
        sc(ws2, 1, c, h, H_FONT_SM, H_FILL)

    for i, ac in enumerate(all_checks, 2):
        ch = ac['checks']
        llm = ac['llm']
        sc(ws2, i, 1, ac['case'])

        # total math
        status = ch['total_math']
        icon = '✅' if status in ('exact', 'close') else ('❌' if status == 'wrong' else '—')
        cell = sc(ws2, i, 2, icon)
        if status == 'wrong': cell.fill = ERR_FILL
        elif status in ('exact', 'close'): cell.fill = OK_FILL
        sc(ws2, i, 3, ch.get('total_math_detail', ''), Font(size=8))

        # tax math
        status = ch['tax_math']
        icon = '✅' if status in ('exact', 'close') else ('❌' if status == 'wrong' else '—')
        cell = sc(ws2, i, 4, icon)
        if status == 'wrong': cell.fill = ERR_FILL
        elif status in ('exact', 'close'): cell.fill = OK_FILL
        sc(ws2, i, 5, ch.get('tax_math_detail', ''), Font(size=8))

        # Items
        sc(ws2, i, 6, ch['item_total'])
        sc(ws2, i, 7, ch['item_math_ok'])
        wrong_cell = sc(ws2, i, 8, ch['item_math_wrong'])
        if ch['item_math_wrong'] > 0: wrong_cell.fill = ERR_FILL

        # InWord
        iw = ch['inword_range']
        icon = '✅' if iw == 'match' else ('❌' if iw == 'mismatch' else '—')
        cell = sc(ws2, i, 9, icon)
        if iw == 'mismatch': cell.fill = ERR_FILL
        elif iw == 'match': cell.fill = OK_FILL

        # Key amounts
        sc(ws2, i, 10, str(llm.get('totalAmount', ''))[:20], Font(size=8))
        sc(ws2, i, 11, str(llm.get('preTaxPrice', ''))[:20], Font(size=8))
        sc(ws2, i, 12, str(llm.get('taxAmount', ''))[:20], Font(size=8))
        sc(ws2, i, 13, str(llm.get('taxPercent', ''))[:10], Font(size=8))

    ws2.freeze_panes = 'B2'
    ws2.column_dimensions['A'].width = 6
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 35
    ws2.column_dimensions['D'].width = 12
    ws2.column_dimensions['E'].width = 40
    for cl in ['F','G','H']: ws2.column_dimensions[cl].width = 10
    ws2.column_dimensions['I'].width = 12
    for cl in ['J','K','L']: ws2.column_dimensions[cl].width = 16
    ws2.column_dimensions['M'].width = 10

    # ═══════════════════════════════════════════════════════════════════════════
    # Sheet 3: LLM Fields (all field values)
    # ═══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("LLM Fields")
    ws3.sheet_properties.tabColor = "548235"

    headers = ['Case'] + ALL_FIELDS
    for c, h in enumerate(headers, 1):
        sc(ws3, 1, c, h, H_FONT_SM, H_FILL)

    for i, ac in enumerate(all_checks, 2):
        sc(ws3, i, 1, ac['case'])
        for j, f in enumerate(ALL_FIELDS, 2):
            v = ac['llm'].get(f)
            val_str = str(v)[:60] if v is not None else ''
            cell = sc(ws3, i, j, val_str, Font(size=8))
            if v is None:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    ws3.freeze_panes = 'B2'
    ws3.column_dimensions['A'].width = 6

    # ═══════════════════════════════════════════════════════════════════════════
    # Sheet 4: LLM Items (all items)
    # ═══════════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("LLM Items")
    ws4.sheet_properties.tabColor = "548235"

    item_h = ['Case', 'Item#', 'productName', 'quantity', 'unitPrice', 'amount',
              'qty×price', 'Math Check']
    for c, h in enumerate(item_h, 1):
        sc(ws4, 1, c, h, H_FONT_SM, H_FILL)

    row = 2
    for ac in all_checks:
        items = ac['llm'].get('itemList', []) or []
        for idx, item in enumerate(items):
            sc(ws4, row, 1, ac['case'])
            sc(ws4, row, 2, idx + 1)
            sc(ws4, row, 3, str(item.get('productName', ''))[:50], Font(size=8))
            sc(ws4, row, 4, str(item.get('quantity', ''))[:15], Font(size=8))
            sc(ws4, row, 5, str(item.get('unitPrice', ''))[:15], Font(size=8))
            sc(ws4, row, 6, str(item.get('amount', ''))[:15], Font(size=8))

            qty = safe_float(item.get('quantity'))
            price = safe_float(item.get('unitPrice'))
            amt = safe_float(item.get('amount'))
            if qty is not None and price is not None:
                expected = qty * price
                sc(ws4, row, 7, f"{expected:.0f}", Font(size=8))
                if amt is not None:
                    ok = abs(amt - expected) < 1 or (abs(amt) > 0 and abs(amt - expected)/abs(amt) < 0.01)
                    icon = '✅' if ok else '❌'
                    cell = sc(ws4, row, 8, icon)
                    if not ok: cell.fill = ERR_FILL
                    else: cell.fill = OK_FILL
                else:
                    sc(ws4, row, 8, '—')
            else:
                sc(ws4, row, 7, '', Font(size=8))
                sc(ws4, row, 8, '—')
            row += 1

    ws4.freeze_panes = 'C2'
    ws4.column_dimensions['A'].width = 6
    ws4.column_dimensions['B'].width = 6
    ws4.column_dimensions['C'].width = 40
    ws4.column_dimensions['D'].width = 12
    ws4.column_dimensions['E'].width = 14
    ws4.column_dimensions['F'].width = 16
    ws4.column_dimensions['G'].width = 16
    ws4.column_dimensions['H'].width = 10

    # ═══════════════════════════════════════════════════════════════════════════
    wb.save(output_path)
    print(f"\n✅ Report saved: {output_path}")
    print(f"   4 Sheets: Summary, Case Validation, LLM Fields, LLM Items")
    print(f"   Cases: {len(results)}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--type", default="gtgt", choices=["gtgt", "commercial"])
    args = parser.parse_args()

    if args.type == "gtgt":
        import tests.test_cases_gtgt as tc_mod
    else:
        import tests.test_cases_commercial as tc_mod

    output = os.path.join(os.path.dirname(__file__), '..', f'{args.type}_llm_standalone_report.xlsx')
    print(f"Generating LLM Standalone Report ({args.type.upper()})...")
    print("=" * 60)
    case_ids = get_case_ids(tc_mod)
    print(f"  Found {len(case_ids)} test cases")
    results = get_llm_data(tc_mod, case_ids)
    create_excel(results, output, args.type)
