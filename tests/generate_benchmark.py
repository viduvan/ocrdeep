#!/usr/bin/env python3
"""
Generate benchmark Excel files from test_cases_commercial and test_cases_gtgt.
Layout:
  Row 1: Header (blue background, white text)
  Row 2: % FIELD NOT NULL per column (yellow background)
  Row 3+: Data rows (green=populated, red=null)
Columns: T | TÊN FILE | Tổng % | invoiceName | invoiceID | ... | invoiceTotalInWord
"""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), ".."))

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import date as date_type


def normalize_date(val):
    if isinstance(val, date_type):
        return val.strftime("%d/%m/%Y")
    if isinstance(val, str) and "-" in val:
        p = val.split("-")
        if len(p) == 3:
            return f"{p[2]}/{p[1]}/{p[0]}"
    return val


ALL_FIELDS = [
    "invoiceName", "invoiceID", "invoiceDate", "invoiceSerial",
    "invoiceFormNo", "paymentMethod", "currency",
    "sellerName", "sellerTaxCode", "sellerEmail",
    "sellerAddress", "sellerPhoneNumber", "sellerBank", "sellerBankAccountNumber",
    "buyerName", "buyerTaxCode", "buyerEmail",
    "buyerAddress", "buyerPhoneNumber", "buyerBank", "buyerBankAccountNumber",
    "preTaxPrice", "discountTotal", "taxPercent", "taxAmount",
    "totalAmount", "invoiceTotalInWord",
]

# Styles
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
PCT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
PCT_FONT = Font(name="Arial", bold=True, size=10)
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
DATA_FONT = Font(name="Arial", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def generate_benchmark(test_module_name, output_filename, sheet_title):
    """Generate a benchmark Excel file from a test module."""
    mod = __import__(test_module_name)

    from src.parsers.block_invoice_parser import parse_invoice_block_based
    from src.parsers.block_invoice_zoomtext_parser import parse_zoom_header

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    # ── HEADER ROW (row 1) ──
    headers = ["T", "TÊN FILE", "Tổng %"] + ALL_FIELDS + ["itemList"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # ── ROW 2: % FIELD NOT NULL (placeholder, filled after data) ──
    # Will be filled after processing all cases

    # ── DATA ROWS (row 3+) ──
    # Track per-column non-null counts
    field_nonnull_counts = {f: 0 for f in ALL_FIELDS}
    itemlist_nonnull_count = 0
    total_cases = 0

    results = []  # Store (case_id, filename, d) for writing

    for case_id, filename, raw_fn, zoom_fn in mod.CASES:
        raw_text = raw_fn()
        zoom_text = zoom_fn()
        invoice = parse_invoice_block_based(raw_text)
        if zoom_text and zoom_text.strip():
            parse_zoom_header(zoom_text.splitlines(), invoice)
        d = invoice.model_dump()
        if d.get("invoiceDate"):
            d["invoiceDate"] = normalize_date(d["invoiceDate"])
        results.append((case_id, filename, d))
        total_cases += 1

        # Count non-null per field
        for field_name in ALL_FIELDS:
            if d.get(field_name) is not None:
                field_nonnull_counts[field_name] += 1
        items = d.get("itemList") or []
        if len(items) > 0:
            itemlist_nonnull_count += 1

    # Write data rows
    for idx, (case_id, filename, d) in enumerate(results):
        row_idx = idx + 3  # row 3 = first data row

        # T (STT)
        cell = ws.cell(row=row_idx, column=1, value=case_id)
        cell.font = DATA_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER

        # TÊN FILE
        cell = ws.cell(row=row_idx, column=2, value=filename)
        cell.font = Font(name="Arial", size=10, bold=True)
        cell.alignment = LEFT
        cell.border = THIN_BORDER

        # Tổng % (per-row score)
        ok_count = sum(1 for f in ALL_FIELDS if d.get(f) is not None)
        pct = ok_count / len(ALL_FIELDS) * 100
        cell = ws.cell(row=row_idx, column=3, value=f"{pct:.1f}%")
        cell.font = Font(name="Arial", size=10, bold=True)
        cell.alignment = CENTER
        cell.border = THIN_BORDER

        # Field columns
        for field_idx, field_name in enumerate(ALL_FIELDS):
            col_idx = field_idx + 4  # starts at column 4
            val = d.get(field_name)
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = THIN_BORDER
            cell.font = DATA_FONT
            cell.alignment = LEFT

            if val is not None:
                cell.value = str(val) if not isinstance(val, (int, float)) else val
                cell.fill = GREEN_FILL
            else:
                cell.value = "null"
                cell.fill = RED_FILL

        # itemList column
        items = d.get("itemList") or []
        item_col = len(ALL_FIELDS) + 4
        cell = ws.cell(row=row_idx, column=item_col, value=len(items))
        cell.font = DATA_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        cell.fill = GREEN_FILL if len(items) > 0 else RED_FILL

        print(f"  [{case_id:02d}] {filename}: {ok_count}/{len(ALL_FIELDS)} ({pct:.1f}%), {len(items)} items")

    # ── ROW 2: % FIELD NOT NULL ──
    # Column 1: empty
    cell = ws.cell(row=2, column=1)
    cell.fill = PCT_FILL
    cell.border = THIN_BORDER

    # Column 2: label
    cell = ws.cell(row=2, column=2, value="% FIELD NOT NULL")
    cell.fill = PCT_FILL
    cell.font = Font(name="Arial", bold=True, size=10, color="FF0000")
    cell.alignment = LEFT
    cell.border = THIN_BORDER

    # Column 3: overall %
    total_nonnull = sum(field_nonnull_counts.values())
    total_possible = total_cases * len(ALL_FIELDS)
    overall_pct = total_nonnull / total_possible * 100 if total_possible else 0
    cell = ws.cell(row=2, column=3, value=f"{overall_pct:.1f}%")
    cell.fill = PCT_FILL
    cell.font = PCT_FONT
    cell.alignment = CENTER
    cell.border = THIN_BORDER

    # Per-field % columns
    for field_idx, field_name in enumerate(ALL_FIELDS):
        col_idx = field_idx + 4
        pct = field_nonnull_counts[field_name] / total_cases * 100 if total_cases else 0
        cell = ws.cell(row=2, column=col_idx, value=f"{pct:.1f}%")
        cell.fill = PCT_FILL
        cell.font = PCT_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # itemList % column
    item_col = len(ALL_FIELDS) + 4
    item_pct = itemlist_nonnull_count / total_cases * 100 if total_cases else 0
    cell = ws.cell(row=2, column=item_col, value=f"{item_pct:.1f}%")
    cell.fill = PCT_FILL
    cell.font = PCT_FONT
    cell.alignment = CENTER
    cell.border = THIN_BORDER

    # ── COLUMN WIDTHS ──
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 42
    ws.column_dimensions['C'].width = 9
    for col_idx in range(4, len(headers) + 1):
        ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = 18

    # ── FREEZE PANES (freeze row 1-2 and columns A-C) ──
    ws.freeze_panes = "D3"

    # Save
    output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", output_filename)
    wb.save(output_path)
    print(f"\n✅ Saved: {output_path}")
    print(f"   {total_nonnull}/{total_possible} fields ({overall_pct:.1f}%)")
    return output_path


if __name__ == "__main__":
    # Generate Commercial benchmark
    print("=" * 60)
    print("Generating Commercial Invoice benchmark...")
    print("=" * 60)
    generate_benchmark(
        "test_cases_commercial",
        "benchmark_commercial.xlsx",
        "Commercial Invoice"
    )

    # Generate GTGT benchmark
    print("\n" + "=" * 60)
    print("Generating GTGT Invoice benchmark...")
    print("=" * 60)
    generate_benchmark(
        "test_cases_gtgt",
        "benchmark_gtgt.xlsx",
        "Invoice GTGT"
    )
